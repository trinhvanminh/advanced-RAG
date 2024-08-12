import csv
import json
import os
from enum import Enum
from typing import List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from rich import print


class OutputType(Enum):
    CSV = 'csv'
    JSON = 'json'


TRANSPOSE_COLUMN = 'Features'
TRANSPOSE_SHEET = 'Strata-Loans'
NORMAL_COLUMN = 'Lender'


def replace_bullet(df: pd.DataFrame) -> pd.DataFrame:
    """ Function to replace \u2022, with * in a DataFrame"""
    return df.map(lambda x: x.replace('\u2022', '*') if isinstance(x, str) else x)


def drop_unnamed(df: pd.DataFrame) -> pd.DataFrame:
    return df.loc[:, ~df.columns.str.contains('^Unnamed')]


def clean_file_name(raw_file_name: str) -> str:
    return (raw_file_name
            .strip()
            .lower()
            .replace(' ', '_')
            .replace(',', '-')
            .replace('&', 'and')
            )


def should_transpose_df(df: pd.DataFrame, sheet_name: str) -> bool:
    return df.columns[0] == TRANSPOSE_COLUMN or sheet_name == TRANSPOSE_SHEET


def transpose_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.T
    new_header = df.iloc[0]
    df = df[1:]
    df.columns = new_header
    df = df.reset_index()
    df = df.rename(columns={'index': NORMAL_COLUMN})
    return df


def is_normal_column(df: pd.DataFrame) -> bool:
    return df.columns[0] == NORMAL_COLUMN


def get_hyperlinks(df: pd.DataFrame, sheet: Worksheet, col_idx: int, should_transpose: bool) -> list[str]:
    hyperlinks = []
    # Start from 2 to skip header
    for row_idx in range(2, sheet.max_row + 1):
        if should_transpose:
            cell = sheet.cell(row=col_idx, column=row_idx)
        else:
            cell = sheet.cell(row=row_idx, column=col_idx)

        hyperlink = cell.hyperlink.target if cell.hyperlink else cell.value or ''
        hyperlinks.append(hyperlink)

        if row_idx == len(df.index) + 1:
            break

    return hyperlinks


def add_hyperlinks(df: pd.DataFrame, sheet: Worksheet, should_transpose: bool):
    for col_idx, col in enumerate(df.columns, start=1):
        if 'website' in col.lower() or 'link' in col.lower():
            hyperlinks = get_hyperlinks(df, sheet, col_idx, should_transpose)
            df[col] = hyperlinks


def save_data(df: pd.DataFrame, output_folder_path: str, output_file_path: str, output_type: OutputType) -> bool:
    try:
        if output_type == OutputType.CSV:
            if not os.path.exists(output_folder_path):
                os.makedirs(output_folder_path)

            df.to_csv(
                output_file_path,
                index=False,
                # sep=',',
                encoding="utf-8",
                quoting=csv.QUOTE_ALL
            )

        elif output_type == OutputType.JSON:
            # Convert the transposed DataFrame to a dictionary
            data_dict = df.to_dict(orient='records')

            # Convert the dictionary to a JSON string
            json_data = json.dumps(data_dict, indent=4)

            with open(output_file_path, 'w') as json_file:
                json_file.write(json_data)

        return True
    except Exception as e:
        print(f"Error saving {output_file_path} file: {e}")
        return False


def load_excel_file(
    file_path: str = 'data/raw/Broker Matrix as at 040724.xlsx',
    output_type: OutputType = OutputType.CSV,
    output_folder: str = './data/preprocessed',
    overwrite: bool = False,
):
    unstructured_sheets: List[str] = []
    output_folder_path = f'{output_folder}/{output_type.value}'

    # Load the workbook using openpyxl to get sheet names
    workbook = load_workbook(file_path)
    sheet_names = workbook.sheetnames

    # Loop through all sheets
    for sheet_name in sheet_names:
        # Save the JSON data to a file named after the sheet
        file_name = clean_file_name(sheet_name)
        output_file_path = f'{output_folder_path}/{file_name}_data.{output_type.value}'

        if not overwrite and os.path.exists(output_file_path):
            continue

        # Read the sheet into a DataFrame using pandas
        df = pd.read_excel(file_path, sheet_name=sheet_name).fillna("")

        # Transpose the DataFrame
        should_transpose = should_transpose_df(df, sheet_name)

        if should_transpose:
            df = transpose_df(df)

        elif not is_normal_column(df):
            unstructured_sheets.append(sheet_name)
            continue

        sheet = workbook[sheet_name]
        add_hyperlinks(df, sheet, should_transpose)

        # Drop all unnamed columns
        df = drop_unnamed(df)

        # Strip header names
        df.columns = df.columns.str.strip()

        # Replace any occurrences of \u2022 with a dash (*) in the transposed DataFrame
        df = replace_bullet(df)

        save_status = save_data(df,
                                output_folder_path,
                                output_file_path,
                                output_type
                                )
        if save_status:
            print((f"Data for sheet '{sheet_name}' "
                   f"has been converted to '{output_type.value}' "
                   f"and saved to '{output_file_path}'."
                   ))

    print(f"All sheets have been processed and saved as {output_type.value}.")
    if len(unstructured_sheets) > 0:
        print("Except these unstructured sheets:")
        print(unstructured_sheets)


if __name__ == '__main__':
    # Load the Excel file
    excel_file_path = 'data/raw/Broker Matrix as at 040724.xlsx'
    output_type: OutputType = OutputType.CSV
    output_folder: str = './data/preprocessed'

    load_excel_file(
        file_path=excel_file_path,
        output_type=output_type,
        output_folder=output_folder,
        overwrite=False
    )
